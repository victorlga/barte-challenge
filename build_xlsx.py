#!/usr/bin/env python3
"""Build Nuvio_Tech_Corrigida.xlsx from Nuvio_Tech_Base_Financeira.xlsx"""
import openpyxl
import re
import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = '/Users/victor/Playground/barte-challenge/Nuvio_Tech_Base_Financeira.xlsx'
OUT = '/Users/victor/Playground/barte-challenge/Nuvio_Tech_Corrigida.xlsx'

src = openpyxl.load_workbook(SRC, data_only=True)

# ── STYLES ─────────────────────────────────────────────────────────────────
H_FILL   = PatternFill("solid", fgColor="1F3864")
H_FONT   = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL = PatternFill("solid", fgColor="EAF2FF")
FLAG_FILL= PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFCCCC")
GRN_FILL = PatternFill("solid", fgColor="CCFFCC")
GRY_FILL = PatternFill("solid", fgColor="F2F2F2")
BOLD     = Font(bold=True)
BOLD_RED = Font(bold=True, color="CC0000")
NUM_FMT  = '#,##0.00'
DATE_FMT = 'DD/MM/YYYY'
PCT_FMT  = '0.0%'

MES_PT = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
          7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}

def mes_label(dt):
    if not isinstance(dt, datetime.datetime): return ''
    return f"{MES_PT[dt.month]}/{str(dt.year)[2:]}"

def hdr(cell, txt=None):
    if txt is not None: cell.value = txt
    cell.fill = H_FILL; cell.font = H_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

def alt(ws, row, ncols):
    if row % 2 == 0:
        for c in range(1, ncols+1):
            ws.cell(row, c).fill = ALT_FILL

# ── EXTRATO CLASSIFICATION ─────────────────────────────────────────────────
def cat_extrato(desc):
    if not desc: return 'Outros'
    d = desc.upper()
    if 'INVESTIDOR' in d and 'APORTE' in d:       return 'Aporte de Investidor'
    if re.search(r'CLIENTE\s', d):                  return 'Receita de Cliente'
    if re.search(r'SALÁRI|FOLHA', d):               return 'Pessoal'
    if re.search(r'\bAWS\b|GOOGLE CLOUD', d):       return 'Infraestrutura Cloud'
    if re.search(r'SLACK|GITHUB|HUBSPOT|NOTION|DATADOG', d): return 'SaaS/Ferramentas'
    if 'PLANO SA' in d:                             return 'Benefícios'
    if 'ALUGUEL' in d:                              return 'Aluguel e Facilities'
    if re.search(r'IMPOSTOS|DAS', d):               return 'Impostos'
    if 'ENERGIA' in d:                              return 'Utilities'
    if 'SEGURO' in d:                               return 'Seguros'
    if 'VIAGEM' in d:                               return 'Viagens'
    if re.search(r'FREELANCER|FORNECEDOR', d):      return 'Freelancers/Fornecedores'
    if 'TRANSFER' in d and 'RESERVA' in d:          return 'Transferência Interna'
    if 'RENDIMENTO' in d:                           return 'Receita Financeira'
    if 'ESTORNO' in d:                              return 'Estorno'
    return 'Outros'

def operacional(cat):
    return 'Não' if cat in ('Transferência Interna','Receita Financeira',
                             'Estorno','Aporte de Investidor') else 'Sim'

# ── ERP CLASSIFICATION ────────────────────────────────────────────────────
FORN_MAP = [
    ('AWS',          'Infraestrutura Cloud'),
    ('GOOGLE CLOUD', 'Infraestrutura Cloud'),
    ('GITHUB',       'SaaS/Ferramentas'),
    ('SLACK',        'SaaS/Ferramentas'),
    ('HUBSPOT',      'SaaS/Ferramentas'),
    ('NOTION',       'SaaS/Ferramentas'),
    ('DATADOG',      'SaaS/Ferramentas'),
    ('CONTAAZUL',    'SaaS/Ferramentas'),
    ('WEWORK',       'Aluguel e Facilities'),
    ('PORTO SEGURO', 'Seguros'),
    ('SILVA',        'Jurídico e Contabilidade'),
    ('UNIMED',       'Benefícios'),
    ('FREELANCER',   'Freelancers/PJ'),
]

def cat_erp(fornecedor, tipo, cat_orig):
    if tipo == 'Receita': return cat_orig
    if not fornecedor: return cat_orig
    fu = fornecedor.upper()
    for k, v in FORN_MAP:
        if k in fu: return v
    return cat_orig

# ── NF DUPLICADA DETECTION ────────────────────────────────────────────────
def find_dup_nfs(erp_rows):
    """Return set of NF/Doc that appear with >1 distinct client"""
    from collections import defaultdict
    nf_clients = defaultdict(set)
    for r in erp_rows:
        nf, cliente = r[7], r[9]
        if nf and cliente:
            nf_clients[nf].add(cliente)
    return {nf for nf, clients in nf_clients.items() if len(clients) > 1}

# ══════════════════════════════════════════════════════════════════════════
# TAB 1 — EXTRATO_LIMPO
# ══════════════════════════════════════════════════════════════════════════
def build_extrato_limpo(wb, src):
    ws = wb.create_sheet('Extrato_Limpo')
    heads = ['Data','Descrição','Tipo','Valor (R$)','Ref. Interna','Banco',
             'Categoria','Operacional','Mês']
    for c, h in enumerate(heads, 1): hdr(ws.cell(1, c), h)

    row = 2
    for r in src['Extrato_Bancario'].iter_rows(min_row=3, values_only=True):
        data, desc, tipo, valor, saldo, ref_int, banco = r
        if data is None: continue
        cat = cat_extrato(desc)
        op  = operacional(cat)
        mes = mes_label(data)

        ws.cell(row,1,data).number_format = DATE_FMT
        ws.cell(row,2,desc)
        ws.cell(row,3,tipo)
        ws.cell(row,4,valor).number_format = NUM_FMT
        ws.cell(row,5,ref_int)
        ws.cell(row,6,banco)
        c7 = ws.cell(row,7,cat)
        ws.cell(row,8,op)
        ws.cell(row,9,mes)

        if op == 'Não': c7.fill = FLAG_FILL
        alt(ws, row, 9) if op == 'Sim' else None
        row += 1

    for c,w in zip(range(1,10),[12,42,18,14,14,7,26,12,8]):
        cw(ws, c, w)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:I{row-1}'
    print(f'  Extrato_Limpo: {row-2} rows')
    return row - 2  # last data row index (1-based offset)

# ══════════════════════════════════════════════════════════════════════════
# TAB 2 — ERP_LIMPO
# ══════════════════════════════════════════════════════════════════════════
def build_erp_limpo(wb, src):
    ws = wb.create_sheet('ERP_Limpo')
    heads = ['Data Lançamento','Data Competência','Descrição','Centro de Custo',
             'Categoria (Original)','Valor (R$)','Tipo','NF/Doc','Status',
             'Fornecedor/Cliente','Categoria_Corrigida','Flag_Competencia',
             'Flag_NF_Duplicada','Flag_Sem_Doc','Mês_Competencia']
    for c, h in enumerate(heads, 1): hdr(ws.cell(1, c), h)

    erp_rows = list(src['Lancamentos_ERP'].iter_rows(min_row=3, values_only=True))
    erp_rows = [r for r in erp_rows if r[0] is not None]
    dup_nfs = find_dup_nfs(erp_rows)

    row = 2
    for r in erp_rows:
        dt_lanc, dt_comp, desc, cc, cat_orig, valor, tipo, nf, status, forn = r
        cat_corr = cat_erp(forn, tipo, cat_orig)

        # Flag: competência diverge do mês de lançamento
        flag_comp = ''
        if isinstance(dt_lanc, datetime.datetime) and isinstance(dt_comp, datetime.datetime):
            if dt_lanc.month != dt_comp.month or dt_lanc.year != dt_comp.year:
                flag_comp = 'DIVERGENTE'

        flag_nf = 'DUPLICADA' if (nf and nf in dup_nfs) else ''
        flag_doc = 'SEM DOCUMENTO' if (tipo == 'Despesa' and not nf) else ''
        mes_comp = mes_label(dt_comp) if isinstance(dt_comp, datetime.datetime) else ''

        ws.cell(row,1,dt_lanc).number_format  = DATE_FMT
        ws.cell(row,2,dt_comp).number_format  = DATE_FMT
        ws.cell(row,3,desc)
        ws.cell(row,4,cc)
        ws.cell(row,5,cat_orig)
        ws.cell(row,6,valor).number_format    = NUM_FMT
        ws.cell(row,7,tipo)
        ws.cell(row,8,nf)
        ws.cell(row,9,status)
        ws.cell(row,10,forn)
        ws.cell(row,11,cat_corr)

        c12 = ws.cell(row,12,flag_comp)
        c13 = ws.cell(row,13,flag_nf)
        c14 = ws.cell(row,14,flag_doc)
        ws.cell(row,15,mes_comp)

        if flag_comp: c12.fill = FLAG_FILL
        if flag_nf:   c13.fill = RED_FILL
        if flag_doc:  c14.fill = RED_FILL
        alt(ws, row, 15)
        row += 1

    for c,w in zip(range(1,16),
                   [12,12,38,14,22,14,9,12,12,20,22,14,16,14,10]):
        cw(ws, c, w)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:O{row-1}'
    print(f'  ERP_Limpo: {row-2} rows, {len(dup_nfs)} NFs duplicadas')

# ══════════════════════════════════════════════════════════════════════════
# TAB 3 — RECEITA_COMPLETA
# ══════════════════════════════════════════════════════════════════════════
def build_receita_completa(wb, src):
    ws = wb.create_sheet('Receita_Completa')
    src_rc = src['Receita_Clientes']

    # ── Seção 1: MRR original ─────────────────────────────────────────
    ws.cell(1,1,'SEÇÃO 1 — MRR POR CLIENTE (fonte: Receita_Clientes original)').font = BOLD
    ws.merge_cells('A1:J1')

    rc_rows = list(src_rc.iter_rows(min_row=3, values_only=True))  # header at row 3, data 4+
    # row 3 is the header row in the source
    sec1_head = rc_rows[0]  # ('Cliente', 'Plano', ...)
    for c, h in enumerate(sec1_head, 1):
        hdr(ws.cell(2, c), h)
    data_row = 3
    for r in rc_rows[1:]:
        if r[0] is None: continue
        for c, v in enumerate(r, 1):
            cell = ws.cell(data_row, c, v)
            if isinstance(v, (int, float)) and c >= 3:
                cell.number_format = NUM_FMT
        alt(ws, data_row, 10)
        data_row += 1

    # Total row with formulas
    ws.cell(data_row, 1, 'TOTAL').font = BOLD
    for col, letter in [(3,'C'),(4,'D'),(5,'E'),(6,'F')]:
        ws.cell(data_row, col,
                f'=SUM({letter}3:{letter}{data_row-1})').number_format = NUM_FMT
        ws.cell(data_row, col).font = BOLD

    # ── Seção 2: Receita Total por Cliente ────────────────────────────
    sec2_row = data_row + 3
    ws.cell(sec2_row, 1,
            'SEÇÃO 2 — RECEITA TOTAL POR CLIENTE (Q4 2025)').font = BOLD
    ws.merge_cells(f'A{sec2_row}:F{sec2_row}')

    s2h = ['Cliente','MRR Q4 (Faturado)','Receita ERP Q4',
           'Receita Extrato Q4','Delta ERP vs Extrato','Obs.']
    for c, h in enumerate(s2h, 1): hdr(ws.cell(sec2_row+1, c), h)

    clients = [
        ('Alpha Tech',    '*ALPHA TECH*'),
        ('Beta SaaS',     '*BETA SAAS*'),
        ('Gamma Digital', '*GAMMA DIGITAL*'),
        ('Delta Cloud',   '*DELTA CLOUD*'),
        ('Epsilon Data',  '*EPSILON DATA*'),
        ('Zeta AI',       '*ZETA AI*'),
        ('Theta Labs',    '*THETA LABS*'),
        ('Iota Systems',  '*IOTA SYSTEMS*'),
        ('Kappa Pay',     '*KAPPA PAY*'),
        ('Lambda Fin',    None),
        ('Mu Analytics',  None),
        ('Nu Robotics',   None),
        ('Omicron Health',None),
        ('Pi Logistics',  None),
        ('Rho Security',  None),
    ]

    # MRR Q4 per client from Seção 1 (D col = Out, E = Nov, F = Dez; data rows 3..data_row-1)
    # We'll reference Receita_Clientes columns by matching client name in Seção 1
    # Simpler: compute MRR sums from source data
    mrr_map = {}
    for r in rc_rows[1:]:
        if r[0]:
            out_ = r[3] if isinstance(r[3],(int,float)) else 0
            nov_ = r[4] if isinstance(r[4],(int,float)) else 0
            dez_ = r[5] if isinstance(r[5],(int,float)) else 0
            mrr_map[r[0]] = out_ + nov_ + dez_

    dr = sec2_row + 2
    for client, pat in clients:
        mrr_q4 = mrr_map.get(client, 0)
        erp_formula = (f'=SUMIFS(ERP_Limpo!F:F,ERP_Limpo!G:G,"Receita",'
                        f'ERP_Limpo!J:J,"{client}")')
        if pat:
            ext_formula = (f'=SUMIFS(Extrato_Limpo!D:D,'
                           f'Extrato_Limpo!G:G,"Receita de Cliente",'
                           f'Extrato_Limpo!B:B,"{pat}")')
            obs = ''
        else:
            ext_formula = 0
            obs = 'Sem recebimento no extrato Q4 — possível inadimplência'

        ws.cell(dr, 1, client)
        ws.cell(dr, 2, mrr_q4).number_format = NUM_FMT
        ws.cell(dr, 3, erp_formula).number_format = NUM_FMT
        ws.cell(dr, 4, ext_formula).number_format = NUM_FMT
        ws.cell(dr, 5, f'=C{dr}-D{dr}').number_format = NUM_FMT
        c6 = ws.cell(dr, 6, obs)
        if not pat: c6.fill = FLAG_FILL
        alt(ws, dr, 6)
        dr += 1

    # Total row
    ws.cell(dr,1,'TOTAL').font = BOLD
    for c,col in [(2,'B'),(3,'C'),(4,'D')]:
        ws.cell(dr,c,f'=SUM({col}{sec2_row+2}:{col}{dr-1})').number_format = NUM_FMT
        ws.cell(dr,c).font = BOLD

    # ── Seção 3: Clientes sem recebimento ─────────────────────────────
    sec3_row = dr + 3
    ws.cell(sec3_row,1,'SEÇÃO 3 — CLIENTES SEM RECEBIMENTO NO EXTRATO Q4').font = BOLD
    ws.merge_cells(f'A{sec3_row}:D{sec3_row}')
    ws.cell(sec3_row,1).fill = FLAG_FILL
    ws.cell(sec3_row+1,1,'Cliente').font = BOLD
    ws.cell(sec3_row+1,2,'MRR Q4 Faturado').font = BOLD
    ws.cell(sec3_row+1,3,'Risco Churn').font = BOLD
    ws.cell(sec3_row+1,4,'Observação').font = BOLD

    no_extrato = [
        ('Lambda Fin',    mrr_map.get('Lambda Fin',0),    'Médio', 'Nenhum recebimento no extrato Q4'),
        ('Mu Analytics',  mrr_map.get('Mu Analytics',0),  'Baixo', 'Nenhum recebimento no extrato Q4'),
        ('Nu Robotics',   mrr_map.get('Nu Robotics',0),   'Baixo', 'Nenhum recebimento no extrato Q4'),
        ('Omicron Health',mrr_map.get('Omicron Health',0),'Baixo', 'Nenhum recebimento no extrato Q4'),
        ('Pi Logistics',  mrr_map.get('Pi Logistics',0),  'Alto',  'Nenhum recebimento no extrato Q4'),
        ('Rho Security',  mrr_map.get('Rho Security',0),  'Baixo', 'NF pendente/atraso nos 3 meses'),
    ]
    for i, (cl, mrr, churn, obs) in enumerate(no_extrato):
        r = sec3_row + 2 + i
        ws.cell(r,1,cl); ws.cell(r,2,mrr).number_format = NUM_FMT
        ws.cell(r,3,churn); ws.cell(r,4,obs)
        for c in range(1,5): ws.cell(r,c).fill = FLAG_FILL

    for c,w in zip(range(1,7),[20,18,18,18,18,42]):
        cw(ws, c, w)
    ws.freeze_panes = 'A3'
    print(f'  Receita_Completa: OK')

# ══════════════════════════════════════════════════════════════════════════
# TAB 4 — DRE_CORRIGIDA
# ══════════════════════════════════════════════════════════════════════════
def build_dre_corrigida(wb):
    ws = wb.create_sheet('DRE_Corrigida')

    # Header
    ws.cell(1,1,'DRE CORRIGIDA — NUVIO TECH (Q4 2025)').font = BOLD
    ws.merge_cells('A1:G1')
    ws.cell(1,1).fill = H_FILL; ws.cell(1,1).font = H_FONT
    ws.cell(1,1).alignment = Alignment(horizontal='center')

    # Column headers
    for c, txt in enumerate(['','Out/25','Nov/25','Dez/25','Q4 2025',
                              'DRE Original (Out)','Fonte/Nota'], 1):
        hdr(ws.cell(2,c), txt)

    MESES = [('Out/25','B'),('Nov/25','C'),('Dez/25','D')]

    def erp_sum(cat_col, cat_val, mes, tipo='Despesa'):
        """SUMIFS formula on ERP_Limpo"""
        return (f'=SUMIFS(ERP_Limpo!F:F,'
                f'ERP_Limpo!G:G,"{tipo}",'
                f'ERP_Limpo!{cat_col}:{cat_col},"{cat_val}",'
                f'ERP_Limpo!O:O,"{mes}")')

    def erp_sum_receita(mes):
        return (f'=SUMIFS(ERP_Limpo!F:F,'
                f'ERP_Limpo!G:G,"Receita",'
                f'ERP_Limpo!O:O,"{mes}")')

    # --- build rows list: (label, [out_formula, nov_formula, dez_formula], orig_val, fonte, bold, fill)
    # orig_val: value from DRE_Gerencial for comparison (Out/25 col)
    rows_def = []

    def R(label, formulas, orig=None, fonte='', bold=False, fill=None, fmt=NUM_FMT):
        rows_def.append((label, formulas, orig, fonte, bold, fill, fmt))

    def SEP(): rows_def.append(None)  # separator/blank row

    def mk(fn_out, fn_nov, fn_dez): return [fn_out, fn_nov, fn_dez]

    # Receita Bruta
    R('Receita Bruta',
      mk(*[erp_sum_receita(m) for m,_ in MESES]),
      442000, 'ERP_Limpo (todas receitas por Data Competência no Q4)', bold=True)

    # Impostos — ERP não tem dados confiáveis; usar 7.5% estimativa
    R('  (-) Impostos sobre Receita (est. 7,5%)',
      mk(*[f'=-0.075*B{3+0}',f'=-0.075*C{3+0}',f'=-0.075*D{3+0}']),  # placeholder; fix after
      -33150, 'Estimativa: 7,5% Receita Bruta (ERP sem dados de impostos confiáveis)')

    R('  (-) Deduções e Devoluções',
      ['=0','=0','=0'],
      -5000, 'Sem dados ERP — manter 0 ou ajustar manualmente')

    R('(=) Receita Líquida', None, None, '', bold=True)
    SEP()

    R('Custos dos Serviços (COGS)', None, None, '', bold=True)
    R('  Infraestrutura Cloud (AWS+GCP)',
      mk(*[erp_sum('K','Infraestrutura Cloud',m) for m,_ in MESES]),
      -78000, 'ERP_Limpo, Categoria_Corrigida=Infraestrutura Cloud')
    R('  SaaS/Ferramentas (operação)',
      mk(*[erp_sum('K','SaaS/Ferramentas',m) for m,_ in MESES]),
      -12000, 'ERP_Limpo, Categoria_Corrigida=SaaS/Ferramentas')
    R('(=) Lucro Bruto', None, None, '', bold=True)
    R('  Margem Bruta %', None, None, '', fmt=PCT_FMT)
    SEP()

    R('Despesas Operacionais', None, None, '', bold=True)
    def ext_sum(cat_val, mes):
        """SUMIFS formula on Extrato_Limpo for expense categories"""
        return (f'=SUMIFS(Extrato_Limpo!D:D,'
                f'Extrato_Limpo!G:G,"{cat_val}",'
                f'Extrato_Limpo!I:I,"{mes}")')

    R('  Pessoal (Salários + Encargos)',
      mk(*[ext_sum('Pessoal',m) for m,_ in MESES]),
      -185000, 'Extrato_Limpo (SALÁRIOS E ENCARGOS) — ERP sem dados Pessoal',
      fill=FLAG_FILL)
    R('  Benefícios',
      mk(*[erp_sum('K','Benefícios',m) for m,_ in MESES]),
      -28000, 'ERP_Limpo, Categoria_Corrigida=Benefícios')
    R('  Freelancers / PJ',
      mk(*[erp_sum('K','Freelancers/PJ',m) for m,_ in MESES]),
      -22000, 'ERP_Limpo, Categoria_Corrigida=Freelancers/PJ')
    R('  Aluguel e Facilities',
      mk(*[erp_sum('K','Aluguel e Facilities',m) for m,_ in MESES]),
      -15000, 'ERP_Limpo, Categoria_Corrigida=Aluguel e Facilities')
    R('  Jurídico e Contabilidade',
      mk(*[erp_sum('K','Jurídico e Contabilidade',m) for m,_ in MESES]),
      -8000, 'ERP_Limpo, Categoria_Corrigida=Jurídico e Contabilidade')
    R('  Viagens',
      mk(*[erp_sum('K','Viagens',m) for m,_ in MESES]),
      -5000, 'ERP_Limpo, Categoria_Corrigida=Viagens')
    R('  Seguros',
      mk(*[erp_sum('K','Seguros',m) for m,_ in MESES]),
      -4000, 'ERP_Limpo, Categoria_Corrigida=Seguros')
    R('  Outros / Diversos',
      mk(*[erp_sum('K','Outros',m) for m,_ in MESES]),
      -4000, 'ERP_Limpo, Categoria_Corrigida=Outros (categorias não mapeadas)')
    R('(=) Total Despesas Operacionais', None, None, '', bold=True)
    SEP()

    R('(=) EBITDA', None, None, '', bold=True)
    R('  Margem EBITDA %', None, None, '', fmt=PCT_FMT)
    SEP()

    R('  Depreciação e Amortização', [-3000,-3000,-3000],
      -3000, 'DRE original — sem fonte raw melhor')
    R('(=) EBIT', None, None, '', bold=True)
    SEP()

    R('  Receitas Financeiras', [2500,3200,2800],
      2500, 'DRE original — sem fonte raw melhor')
    R('  Despesas Financeiras', [-1500,-1800,-2200],
      -1500, 'DRE original — sem fonte raw melhor')
    R('(=) Resultado Financeiro', None, None, '', bold=True)
    SEP()

    R('(=) Resultado Antes IR/CS', None, None, '', bold=True)
    R('  IR/CS', [0,0,0], 0, '')
    R('(=) Resultado Líquido', None, None, '', bold=True)
    R('  Margem Líquida %', None, None, '', fmt=PCT_FMT)

    # Now write rows and compute subtotal references
    # Track row numbers for key lines
    row_map = {}
    excel_row = 3

    def col(c): return get_column_letter(c)  # 2→B, 3→C, 4→D

    for item in rows_def:
        if item is None:
            excel_row += 1
            continue
        label, formulas, orig, fonte, bold, fill, fmt = item
        ws.cell(excel_row, 1, label)
        if bold: ws.cell(excel_row, 1).font = BOLD
        if fill: ws.cell(excel_row, 1).fill = fill

        row_map[label.strip()] = excel_row

        if formulas:
            for ci, f in enumerate(formulas, 2):  # cols B,C,D
                c = ws.cell(excel_row, ci, f)
                c.number_format = fmt
                if bold: c.font = BOLD

        ws.cell(excel_row, 6, orig).number_format = NUM_FMT if orig else ''
        ws.cell(excel_row, 7, fonte)
        ws.cell(excel_row, 7).alignment = Alignment(wrap_text=True)
        excel_row += 1

    # Fix Impostos rows to reference Receita Bruta row
    rb_row = row_map.get('Receita Bruta', 3)
    imp_row = row_map.get('(-) Impostos sobre Receita (est. 7,5%)', 4)
    if imp_row:
        ws.cell(imp_row,2,f'=-0.075*B{rb_row}').number_format = NUM_FMT
        ws.cell(imp_row,3,f'=-0.075*C{rb_row}').number_format = NUM_FMT
        ws.cell(imp_row,4,f'=-0.075*D{rb_row}').number_format = NUM_FMT

    # Fill computed subtotals
    def fill_subtotal(label, addends, is_pct=False, pct_base_label=None):
        r = row_map.get(label.strip())
        if not r: return
        for ci in [2,3,4]:
            cl = col(ci)
            if is_pct:
                base_r = row_map.get(pct_base_label.strip())
                f = f'=IF({cl}{base_r}=0,0,{cl}{r-1}/{cl}{base_r})'
                ws.cell(r,ci,f).number_format = PCT_FMT
            else:
                parts = [f'{col(ci)}{row_map[a.strip()]}' for a in addends if row_map.get(a.strip())]
                ws.cell(r,ci,f'=SUM({",".join(parts)})').number_format = NUM_FMT
            ws.cell(r,ci).font = BOLD

    # Receita Líquida
    fill_subtotal('(=) Receita Líquida',
        ['Receita Bruta','(-) Impostos sobre Receita (est. 7,5%)','(-) Deduções e Devoluções'])

    # Lucro Bruto
    fill_subtotal('(=) Lucro Bruto',
        ['(=) Receita Líquida','Infraestrutura Cloud (AWS+GCP)','SaaS/Ferramentas (operação)'])

    # Margem Bruta
    rl_r = row_map.get('(=) Receita Líquida')
    lb_r = row_map.get('(=) Lucro Bruto')
    mb_r = row_map.get('Margem Bruta %')
    if mb_r and rl_r and lb_r:
        for ci in [2,3,4]:
            cl = col(ci)
            ws.cell(mb_r,ci,f'=IF({cl}{rl_r}=0,0,{cl}{lb_r}/{cl}{rl_r})').number_format=PCT_FMT

    # Total Despesas Op
    desp_labels = ['Pessoal (Salários + Encargos)','Benefícios','Freelancers / PJ',
                   'Aluguel e Facilities','Jurídico e Contabilidade','Viagens','Seguros',
                   'Outros / Diversos']
    fill_subtotal('(=) Total Despesas Operacionais', desp_labels)

    # EBITDA
    fill_subtotal('(=) EBITDA', ['(=) Lucro Bruto','(=) Total Despesas Operacionais'])

    # Margem EBITDA
    ebitda_r = row_map.get('(=) EBITDA')
    mebitda_r = row_map.get('Margem EBITDA %')
    if mebitda_r and rl_r and ebitda_r:
        for ci in [2,3,4]:
            cl = col(ci)
            ws.cell(mebitda_r,ci,f'=IF({cl}{rl_r}=0,0,{cl}{ebitda_r}/{cl}{rl_r})').number_format=PCT_FMT

    # EBIT
    fill_subtotal('(=) EBIT', ['(=) EBITDA','Depreciação e Amortização'])

    # Resultado Financeiro
    fill_subtotal('(=) Resultado Financeiro', ['Receitas Financeiras','Despesas Financeiras'])

    # Resultado Antes IR/CS
    fill_subtotal('(=) Resultado Antes IR/CS', ['(=) EBIT','(=) Resultado Financeiro'])

    # Resultado Líquido
    fill_subtotal('(=) Resultado Líquido', ['(=) Resultado Antes IR/CS','IR/CS'])

    # Margem Líquida
    rl2_r = row_map.get('(=) Resultado Líquido')
    ml_r = row_map.get('Margem Líquida %')
    if ml_r and rl_r and rl2_r:
        for ci in [2,3,4]:
            cl = col(ci)
            ws.cell(ml_r,ci,f'=IF({cl}{rl_r}=0,0,{cl}{rl2_r}/{cl}{rl_r})').number_format=PCT_FMT

    # Q4 total column (E) for numeric rows
    for r in range(3, excel_row):
        v = ws.cell(r,2).value
        if v and str(v).startswith('=') and '=IF(' not in str(v):
            ws.cell(r,5,f'=SUM(B{r}:D{r})').number_format = ws.cell(r,2).number_format
        elif isinstance(v,(int,float)):
            ws.cell(r,5,f'=SUM(B{r}:D{r})').number_format = NUM_FMT

    for c,w in zip(range(1,8),[36,14,14,14,14,16,42]):
        cw(ws, c, w)
    ws.freeze_panes = 'B3'
    ws.row_dimensions[2].height = 28
    print(f'  DRE_Corrigida: OK')

# ══════════════════════════════════════════════════════════════════════════
# TAB 5 — CAIXA_CORRIGIDO
# ══════════════════════════════════════════════════════════════════════════
def build_caixa_corrigido(wb):
    ws = wb.create_sheet('Caixa_Corrigido')
    ws.cell(1,1,'CAIXA CORRIGIDO — NUVIO TECH (Q4 2025) — Fonte: Extrato_Limpo').font = BOLD
    ws.merge_cells('A1:F1')
    ws.cell(1,1).fill = H_FILL; ws.cell(1,1).font = H_FONT
    ws.cell(1,1).alignment = Alignment(horizontal='center')

    for c, txt in enumerate(['','Out/25','Nov/25','Dez/25','Média Mensal',
                              'Caixa Original (Out)'], 1):
        hdr(ws.cell(2,c), txt)

    def ext_sum(sign, op, mes, extra_crit=''):
        """SUMIFS on Extrato_Limpo D col"""
        sign_crit = f'Extrato_Limpo!D:D,"{sign}0",' if sign else ''
        op_crit   = f'Extrato_Limpo!H:H,"{op}",' if op else ''
        mes_crit  = f'Extrato_Limpo!I:I,"{mes}"'
        return (f'=SUMIFS(Extrato_Limpo!D:D,'
                f'{sign_crit}{op_crit}{mes_crit}{extra_crit})')

    def cat_sum(cat, mes):
        return (f'=SUMIFS(Extrato_Limpo!D:D,'
                f'Extrato_Limpo!G:G,"{cat}",'
                f'Extrato_Limpo!I:I,"{mes}")')

    rows = [
        ('Saldo Inicial', None, None, 3200000, 'Derivado do extrato: Saldo implícito antes de Out/25'),
        None,
        ('(+) Entradas Operacionais',
         [ext_sum('>','Sim',m) for m in ('Out/25','Nov/25','Dez/25')],
         420000, 'Extrato: Valor>0, Operacional=Sim'),
        ('(-) Saídas Operacionais',
         [ext_sum('<','Sim',m) for m in ('Out/25','Nov/25','Dez/25')],
         -395000, 'Extrato: Valor<0, Operacional=Sim'),
        ('(=) Fluxo Caixa Operacional', None, None, 25000, ''),
        None,
        ('(+) Aporte de Investidores',
         [cat_sum('Aporte de Investidor',m) for m in ('Out/25','Nov/25','Dez/25')],
         0, 'Extrato: Categoria=Aporte de Investidor (não operacional)'),
        ('(+) Receitas Financeiras',
         [cat_sum('Receita Financeira',m) for m in ('Out/25','Nov/25','Dez/25')],
         2500, 'Extrato: Categoria=Receita Financeira'),
        ('(+) Estornos',
         [cat_sum('Estorno',m) for m in ('Out/25','Nov/25','Dez/25')],
         0, 'Extrato: Categoria=Estorno'),
        ('(info) Transferências Internas',
         [cat_sum('Transferência Interna',m) for m in ('Out/25','Nov/25','Dez/25')],
         0, 'Informativo — não incluso no saldo final'),
        None,
        ('(=) Variação de Caixa', None, None, None, ''),
        ('(=) Saldo Final', None, None, None, ''),
        None,
        ('Burn Rate Mensal (Líquido)', None, None, None, 'Negativo do Fluxo Oper. + Financeiro'),
        ('Runway (meses)', None, None, None,
         'Saldo Final / |Burn Rate|. N/A se burn≤0'),
    ]

    excel_row = 3
    row_map = {}
    for item in rows:
        if item is None:
            excel_row += 1
            continue
        label, formulas, orig, note = item[0], item[1], item[2], item[-1]
        ws.cell(excel_row,1,label)
        row_map[label.strip()] = excel_row

        if formulas:
            for ci, f in enumerate(formulas, 2):
                ws.cell(excel_row,ci,f).number_format = NUM_FMT
        if orig is not None:
            ws.cell(excel_row,5,orig).number_format = NUM_FMT
        ws.cell(excel_row,6,note).alignment = Alignment(wrap_text=True)
        excel_row += 1

    # Saldo Inicial Out: derived from extrato
    si_r = row_map['Saldo Inicial']
    # Saldo implied = first row saldo in extrato - first row valor
    # = 2850000 (as calculated in PROBLEMAS.md)
    ws.cell(si_r,2,2850000).number_format = NUM_FMT
    ws.cell(si_r,2).font = Font(italic=True)  # mark as derived, not formula
    ws.cell(si_r,6,'Saldo implícito derivado do extrato (antes de 01/Out/25). '
                    'Caixa_Runway original usava R$3.200.000 (diferença: R$350k não explicada).')
    # Nov/Dec Saldo Inicial set below after sf_r is determined

    # Fluxo Caixa Operacional
    fco_r = row_map['(=) Fluxo Caixa Operacional']
    ent_r = row_map['(+) Entradas Operacionais']
    sai_r = row_map['(-) Saídas Operacionais']
    for ci in [2,3,4]:
        cl = get_column_letter(ci)
        ws.cell(fco_r,ci,f'={cl}{ent_r}+{cl}{sai_r}').number_format = NUM_FMT
        ws.cell(fco_r,ci).font = BOLD

    # Variação de Caixa
    vc_r = row_map['(=) Variação de Caixa']
    apt_r = row_map['(+) Aporte de Investidores']
    rf_r  = row_map['(+) Receitas Financeiras']
    est_r = row_map['(+) Estornos']
    for ci in [2,3,4]:
        cl = get_column_letter(ci)
        ws.cell(vc_r,ci,
            f'={cl}{fco_r}+{cl}{apt_r}+{cl}{rf_r}+{cl}{est_r}').number_format = NUM_FMT
        ws.cell(vc_r,ci).font = BOLD

    # Saldo Final
    sf_r = row_map['(=) Saldo Final']
    for ci in [2,3,4]:
        cl = get_column_letter(ci)
        if ci == 2:
            ws.cell(sf_r,ci,f'=B{si_r}+B{vc_r}').number_format = NUM_FMT
        else:
            prev_sf = sf_r - 1  # not right; need previous month's saldo final
            # Need to look up previous month saldo final
            # Out→si_r, Nov→B sf_r, Dez→C sf_r
            if ci == 3:
                ws.cell(sf_r,ci,f'=B{sf_r}+C{vc_r}').number_format = NUM_FMT
            else:
                ws.cell(sf_r,ci,f'=C{sf_r}+D{vc_r}').number_format = NUM_FMT
        ws.cell(sf_r,ci).font = BOLD

    # Saldo Inicial Nov/Dec = Saldo Final do mês anterior
    ws.cell(si_r, 3, f'=B{sf_r}').number_format = NUM_FMT
    ws.cell(si_r, 3).font = Font(italic=True)
    ws.cell(si_r, 4, f'=C{sf_r}').number_format = NUM_FMT
    ws.cell(si_r, 4).font = Font(italic=True)

    # Burn Rate
    br_r = row_map['Burn Rate Mensal (Líquido)']
    for ci in [2,3,4]:
        cl = get_column_letter(ci)
        ws.cell(br_r,ci,f'=-{cl}{vc_r}').number_format = NUM_FMT

    # Runway
    rw_r = row_map['Runway (meses)']
    for ci in [2,3,4]:
        cl = get_column_letter(ci)
        ws.cell(rw_r,ci,f'=IF({cl}{br_r}<=0,"N/A",{cl}{sf_r}/{cl}{br_r})').number_format='0.0'

    # Avg column
    ws.cell(si_r,5,'').number_format = NUM_FMT
    for r_label, r_excel in row_map.items():
        vb = ws.cell(r_excel,2).value
        if vb and str(vb).startswith('='):
            ws.cell(r_excel,5,f'=AVERAGE(B{r_excel}:D{r_excel})').number_format = NUM_FMT

    for c,w in zip(range(1,7),[36,14,14,14,14,42]):
        cw(ws, c, w)
    ws.freeze_panes = 'B3'
    print(f'  Caixa_Corrigido: OK')

# ══════════════════════════════════════════════════════════════════════════
# TAB 6 — AP_AR_ANOTADO
# ══════════════════════════════════════════════════════════════════════════
def build_apar_anotado(wb, src):
    ws = wb.create_sheet('AP_AR_Anotado')
    src_ws = src['AP_AR']

    orig_rows = list(src_ws.iter_rows(min_row=3, values_only=True))  # header at row3, data row4+
    head_row = orig_rows[0]
    data_rows = [r for r in orig_rows[1:] if r[0] is not None]

    orig_headers = list(head_row) + ['Ação Recomendada','Prioridade','Recebido_no_Extrato']
    for c, h in enumerate(orig_headers, 1):
        hdr(ws.cell(1, c), h)

    ACOES = {
        'Beta SaaS':        ('Acionar jurídico para cobrança — cliente churnou', 'Alta'),
        'Delta Cloud':      ('Emitir NF corrigida (R$15k) e negociar diferença de R$3k', 'Alta'),
        'Mu Analytics':     ('Enviar NF imediatamente ao cliente', 'Alta'),
        'Alpha Tech':       ('Monitorar — a vencer, NF pendente há 3 meses', 'Média'),
        'Epsilon Data':     ('A vencer — acompanhar recebimento', 'Baixa'),
        'Pi Logistics':     ('Contato comercial urgente + cobrança', 'Alta'),
        'Gamma Digital':    ('Enviar cobrança de setup (nunca enviada)', 'Média'),
        'AWS':              ('Monitorar — aumento 17% vs Out, a vencer', 'Média'),
        'Google Cloud':     ('A vencer — acompanhar fatura', 'Baixa'),
        'WeWork':           ('A vencer — pagamento normal', 'Baixa'),
        'Unimed':           ('A vencer — pagamento normal', 'Baixa'),
        'Freelancer A':     ('Regularizar contrato formal antes de pagar', 'Alta'),
        'Freelancer B':     ('Solicitar NF ao fornecedor antes de pagar', 'Média'),
        'Silva & Advogados':('Revisar contrato e alinhar valor contestado', 'Alta'),
        'Receita Federal':  ('A vencer — garantir pagamento em dia (DAS)', 'Média'),
        'Folha':            ('Pago com atraso — registrar ocorrência RH', 'Baixa'),
    }

    # Client→extrato pattern map (for AR)
    EXT_PAT = {
        'Beta SaaS':     '*BETA SAAS*',
        'Delta Cloud':   '*DELTA CLOUD*',
        'Mu Analytics':  None,
        'Alpha Tech':    '*ALPHA TECH*',
        'Epsilon Data':  '*EPSILON DATA*',
        'Pi Logistics':  None,
        'Gamma Digital': '*GAMMA DIGITAL*',
    }

    row = 2
    for dr in data_rows:
        tipo, cliente, desc, valor, venc, status, dias, cc, obs = dr
        acao, prior = ACOES.get(cliente, ('—', 'Baixa'))

        for c, v in enumerate(dr, 1):
            cell = ws.cell(row, c, v)
            if c == 4: cell.number_format = NUM_FMT
            if c == 5 and isinstance(v, datetime.datetime): cell.number_format = DATE_FMT

        ws.cell(row, 10, acao)
        ws.cell(row, 11, prior)

        # Recebido_no_Extrato: only meaningful for AR
        if tipo == 'AR':
            pat = EXT_PAT.get(cliente)
            if pat:
                f = (f'=SUMIFS(Extrato_Limpo!D:D,'
                     f'Extrato_Limpo!G:G,"Receita de Cliente",'
                     f'Extrato_Limpo!B:B,"{pat}")')
                ws.cell(row, 12, f).number_format = NUM_FMT
            else:
                ws.cell(row, 12, 0).number_format = NUM_FMT

        # Color by status/priority
        if status == 'Vencida' and tipo == 'AR':
            for c in range(1, 13): ws.cell(row,c).fill = RED_FILL
        elif status == 'Vencida' and tipo == 'AP':
            for c in range(1, 13): ws.cell(row,c).fill = FLAG_FILL
        elif prior == 'Alta':
            ws.cell(row,10).fill = FLAG_FILL; ws.cell(row,11).fill = FLAG_FILL
        alt(ws, row, 12)
        row += 1

    for c,w in zip(range(1,13),
                   [5,18,30,14,12,10,10,14,32,36,8,16]):
        cw(ws, c, w)
    ws.freeze_panes = 'A2'
    print(f'  AP_AR_Anotado: OK')

# ══════════════════════════════════════════════════════════════════════════
# TAB 7 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════
def build_dashboard(wb):
    ws = wb.create_sheet('Dashboard')
    ws.sheet_view.showGridLines = False

    def title(row, col, txt, span=4):
        ws.cell(row, col, txt).font = Font(bold=True, size=12, color='FFFFFF')
        ws.cell(row, col).fill = H_FILL
        ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col+span-1)

    def subhdr(row, col, txt):
        ws.cell(row, col, txt).font = Font(bold=True, size=10)
        ws.cell(row, col).fill = GRY_FILL

    # ── Seção 1: Receita Real vs. Reportada ───────────────────────────
    title(1, 1, '1. RECEITA REAL VS. REPORTADA (Q4 2025)', 5)
    for c, h in enumerate(['','DRE Original','DRE Corrigida','Diferença','Ratio'], 1):
        subhdr(2, c, h)

    # Hardcoded DRE Original values (DRE_Gerencial sheet not in this workbook)
    dre_data = [
        ('Receita Bruta Out/25',
         442000,
         '=DRE_Corrigida!B3',
         '=C3-B3', '=IF(B3=0,"N/A",C3/B3)'),
        ('Receita Bruta Nov/25',
         435500,
         '=DRE_Corrigida!C3',
         '=C4-B4', '=IF(B4=0,"N/A",C4/B4)'),
        ('Receita Bruta Dez/25',
         416500,
         '=DRE_Corrigida!D3',
         '=C5-B5', '=IF(B5=0,"N/A",C5/B5)'),
        ('Receita Bruta Q4',
         1294000,
         '=DRE_Corrigida!E3',
         '=C6-B6', '=IF(B6=0,"N/A",C6/B6)'),
    ]
    for i, (label, orig, corr, delta, ratio) in enumerate(dre_data):
        r = 3 + i
        ws.cell(r,1,label)
        for c,v in enumerate([orig,corr,delta,ratio],2):
            cell = ws.cell(r,c,v)
            cell.number_format = NUM_FMT if c < 5 else '0.0x'
        if i % 2 == 0:
            for c in range(1,6): ws.cell(r,c).fill = ALT_FILL

    # ── Seção 2: Saúde do Caixa ────────────────────────────────────────
    title(9, 1, '2. SAÚDE DO CAIXA', 5)
    cash_rows = [
        ('Saldo Inicial Out/25 (extrato)',   '=Caixa_Corrigido!B3',  ''),
        ('Saldo Final Dez/25 (extrato)',     '=Caixa_Corrigido!D15', ''),
        ('Entradas Operacionais Q4',
         '=Caixa_Corrigido!B5+Caixa_Corrigido!C5+Caixa_Corrigido!D5', ''),
        ('Saídas Operacionais Q4',
         '=Caixa_Corrigido!B6+Caixa_Corrigido!C6+Caixa_Corrigido!D6', ''),
        ('Aporte Investidores Q4',
         '=Caixa_Corrigido!B9+Caixa_Corrigido!C9+Caixa_Corrigido!D9', ''),
        ('Burn Rate Médio Mensal',   '=AVERAGE(Caixa_Corrigido!B17:D17)', ''),
        ('Runway (Dez/25, meses)',   '=Caixa_Corrigido!D18', ''),
    ]
    for i, (label, formula, note) in enumerate(cash_rows):
        r = 10 + i
        ws.cell(r,1,label)
        ws.cell(r,2,formula).number_format = NUM_FMT
        ws.cell(r,3,note)
        if i % 2 == 0:
            for c in range(1,4): ws.cell(r,c).fill = ALT_FILL

    # ── Seção 3: Alertas ───────────────────────────────────────────────
    title(19, 1, '3. ALERTAS E FLAGS', 5)
    alert_rows = [
        ('NFs Duplicadas (ERP)',
         '=COUNTIF(ERP_Limpo!M:M,"DUPLICADA")',
         'NFs com mesmo número em clientes diferentes'),
        ('Despesas sem Documento (ERP)',
         '=COUNTIF(ERP_Limpo!N:N,"SEM DOCUMENTO")',
         'Despesas sem NF ou comprovante'),
        ('Lançamentos com Competência Divergente',
         '=COUNTIF(ERP_Limpo!L:L,"DIVERGENTE")',
         'Data competência ≠ mês de lançamento'),
        ('Lançamentos não Conciliados (ERP)',
         '=COUNTIFS(ERP_Limpo!I:I,"<>Conciliado",ERP_Limpo!A:A,"<>")',
         'Status Pendente + Divergente + Em Aberto'),
        ('AR Vencido (R$)',
         '=SUMIFS(AP_AR_Anotado!D:D,AP_AR_Anotado!A:A,"AR",AP_AR_Anotado!F:F,"Vencida")',
         'Contas a receber vencidas'),
        ('AP Vencido (R$)',
         '=SUMIFS(AP_AR_Anotado!D:D,AP_AR_Anotado!A:A,"AP",AP_AR_Anotado!F:F,"Vencida")',
         'Contas a pagar vencidas'),
        ('Clientes sem Recebimento no Extrato Q4',
         6, 'Lambda Fin, Mu Analytics, Nu Robotics, Omicron Health, Pi Logistics, Rho Security'),
    ]
    for i, (label, formula, note) in enumerate(alert_rows):
        r = 20 + i
        ws.cell(r,1,label)
        cell = ws.cell(r,2, formula)
        if isinstance(formula, str) and formula.startswith('='):
            cell.number_format = '#,##0' if 'SUMIFS' in formula else '0'
        ws.cell(r,3,note).alignment = Alignment(wrap_text=True)
        ws.cell(r,2).fill = FLAG_FILL
        if i % 2 == 0:
            ws.cell(r,1).fill = ALT_FILL; ws.cell(r,3).fill = ALT_FILL

    for c,w in zip(range(1,6),[38,18,42,14,10]):
        cw(ws, c, w)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[9].height = 24
    ws.row_dimensions[19].height = 24
    print(f'  Dashboard: OK')

# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

print('Building Nuvio_Tech_Corrigida.xlsx...')
build_extrato_limpo(wb, src)
build_erp_limpo(wb, src)
build_receita_completa(wb, src)
build_dre_corrigida(wb)
build_caixa_corrigido(wb)
build_apar_anotado(wb, src)
build_dashboard(wb)

wb.save(OUT)
print(f'\nSalvo em: {OUT}')
